"""
DAY 16: EXCEL AUTOMATION WITH PYTHON
Combining SQL + Excel for automated reporting

Real-world skill: Generate Excel reports from database queries
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime
import os

class ExcelReportGenerator:
    """Automated Excel report generation from SQL queries"""
    
    def __init__(self, db_name='sales_analysis.db'):
        """Initialize database connection"""
        self.conn = sqlite3.connect(db_name)
        print(f"✅ Connected to: {db_name}\n")
    
    def create_sales_summary_report(self, output_file='Sales_Summary_Report.xlsx'):
        """Generate comprehensive sales summary in Excel"""
        
        print("="*70)
        print("📊 GENERATING SALES SUMMARY REPORT")
        print("="*70)
        
        # Create Excel writer
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # =====================================================================
        # SHEET 1: Executive Summary
        # =====================================================================
        print("\n📄 Sheet 1: Executive Summary...")
        
        query = '''
        SELECT 
            'Total Customers' as metric,
            COUNT(DISTINCT customer_id) as value
        FROM customers
        UNION ALL
        SELECT 
            'Total Orders',
            COUNT(*)
        FROM orders
        UNION ALL
        SELECT 
            'Completed Orders',
            SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END)
        FROM orders
        UNION ALL
        SELECT 
            'Total Revenue',
            ROUND(SUM(CASE WHEN status = 'Completed' THEN total_amount ELSE 0 END), 2)
        FROM orders
        UNION ALL
        SELECT 
            'Average Order Value',
            ROUND(AVG(CASE WHEN status = 'Completed' THEN total_amount END), 2)
        FROM orders;
        '''
        
        df_summary = pd.read_sql_query(query, self.conn)
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # =====================================================================
        # SHEET 2: Top Customers
        # =====================================================================
        print("📄 Sheet 2: Top Customers...")
        
        query = '''
        SELECT 
            c.name as Customer,
            c.city as City,
            c.customer_type as Type,
            COUNT(o.order_id) as Orders,
            ROUND(SUM(o.total_amount), 2) as Total_Spent,
            ROUND(AVG(o.total_amount), 2) as Avg_Order
        FROM customers c
        JOIN orders o ON c.customer_id = o.customer_id
        WHERE o.status = 'Completed'
        GROUP BY c.customer_id, c.name, c.city, c.customer_type
        ORDER BY Total_Spent DESC
        LIMIT 20;
        '''
        
        df_customers = pd.read_sql_query(query, self.conn)
        df_customers.to_excel(writer, sheet_name='Top Customers', index=False)
        
        # =====================================================================
        # SHEET 3: Product Performance
        # =====================================================================
        print("📄 Sheet 3: Product Performance...")
        
        query = '''
        SELECT 
            p.product_name as Product,
            p.category as Category,
            p.price as Unit_Price,
            SUM(o.quantity) as Units_Sold,
            ROUND(SUM(o.total_amount), 2) as Revenue,
            ROUND(AVG(o.total_amount), 2) as Avg_Sale
        FROM products p
        JOIN orders o ON p.product_id = o.product_id
        WHERE o.status = 'Completed'
        GROUP BY p.product_id, p.product_name, p.category, p.price
        ORDER BY Revenue DESC;
        '''
        
        df_products = pd.read_sql_query(query, self.conn)
        df_products.to_excel(writer, sheet_name='Product Performance', index=False)
        
        # =====================================================================
        # SHEET 4: City Analysis
        # =====================================================================
        print("📄 Sheet 4: City Analysis...")
        
        query = '''
        SELECT 
            c.city as City,
            COUNT(DISTINCT c.customer_id) as Customers,
            COUNT(o.order_id) as Orders,
            ROUND(SUM(o.total_amount), 2) as Revenue,
            ROUND(AVG(o.total_amount), 2) as Avg_Order,
            ROUND(SUM(o.total_amount) / COUNT(DISTINCT c.customer_id), 2) as Revenue_Per_Customer
        FROM customers c
        JOIN orders o ON c.customer_id = o.customer_id
        WHERE o.status = 'Completed'
        GROUP BY c.city
        ORDER BY Revenue DESC;
        '''
        
        df_cities = pd.read_sql_query(query, self.conn)
        df_cities.to_excel(writer, sheet_name='City Analysis', index=False)
        
        # =====================================================================
        # SHEET 5: Monthly Trends
        # =====================================================================
        print("📄 Sheet 5: Monthly Trends...")
        
        query = '''
        SELECT 
            strftime('%Y-%m', order_date) as Month,
            COUNT(*) as Orders,
            ROUND(SUM(total_amount), 2) as Revenue,
            ROUND(AVG(total_amount), 2) as Avg_Order
        FROM orders
        WHERE status = 'Completed'
        GROUP BY strftime('%Y-%m', order_date)
        ORDER BY Month;
        '''
        
        df_monthly = pd.read_sql_query(query, self.conn)
        df_monthly.to_excel(writer, sheet_name='Monthly Trends', index=False)
        
        # Save the workbook
        writer.close()
        
        # =====================================================================
        # FORMAT THE EXCEL FILE
        # =====================================================================
        print("\n🎨 Formatting Excel file...")
        self.format_excel_report(output_file)
        
        print(f"\n✅ Report generated: {output_file}")
        print(f"📊 5 sheets created with formatted data and charts")
        
        return output_file
    
    def format_excel_report(self, filename):
        """Add professional formatting to Excel file"""
        
        # Load the workbook
        wb = load_workbook(filename)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells with data
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            # Center align numeric columns
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
        
        # Save formatted workbook
        wb.save(filename)
        print("   ✅ Professional formatting applied")
    
    def create_advanced_report_with_charts(self, output_file='Advanced_Sales_Report.xlsx'):
        """Create report with charts and advanced formatting"""
        
        print("\n" + "="*70)
        print("📊 GENERATING ADVANCED REPORT WITH CHARTS")
        print("="*70)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # =====================================================================
        # SHEET: Top Products Chart
        # =====================================================================
        print("\n📊 Creating Top Products chart...")
        
        ws_products = wb.create_sheet("Top Products")
        
        # Get data
        query = '''
        SELECT 
            p.product_name,
            ROUND(SUM(o.total_amount), 2) as revenue
        FROM products p
        JOIN orders o ON p.product_id = o.product_id
        WHERE o.status = 'Completed'
        GROUP BY p.product_id, p.product_name
        ORDER BY revenue DESC
        LIMIT 10;
        '''
        
        df = pd.read_sql_query(query, self.conn)
        
        # Write headers
        ws_products['A1'] = 'Product'
        ws_products['B1'] = 'Revenue'
        
        # Write data
        for idx, row in df.iterrows():
            ws_products.cell(row=idx+2, column=1, value=row['product_name'])
            ws_products.cell(row=idx+2, column=2, value=row['revenue'])
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Top 10 Products by Revenue"
        chart.x_axis.title = "Product"
        chart.y_axis.title = "Revenue (₹)"
        
        data = Reference(ws_products, min_col=2, min_row=1, max_row=len(df)+1)
        categories = Reference(ws_products, min_col=1, min_row=2, max_row=len(df)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 15
        chart.width = 25
        
        ws_products.add_chart(chart, "D2")
        
        # =====================================================================
        # SHEET: Monthly Trend Chart
        # =====================================================================
        print("📈 Creating Monthly Trends chart...")
        
        ws_monthly = wb.create_sheet("Monthly Trends")
        
        # Get data
        query = '''
        SELECT 
            strftime('%Y-%m', order_date) as month,
            ROUND(SUM(total_amount), 2) as revenue
        FROM orders
        WHERE status = 'Completed'
        GROUP BY strftime('%Y-%m', order_date)
        ORDER BY month;
        '''
        
        df = pd.read_sql_query(query, self.conn)
        
        # Write headers
        ws_monthly['A1'] = 'Month'
        ws_monthly['B1'] = 'Revenue'
        
        # Write data
        for idx, row in df.iterrows():
            ws_monthly.cell(row=idx+2, column=1, value=row['month'])
            ws_monthly.cell(row=idx+2, column=2, value=row['revenue'])
        
        # Create line chart
        chart = LineChart()
        chart.title = "Monthly Revenue Trend"
        chart.x_axis.title = "Month"
        chart.y_axis.title = "Revenue (₹)"
        
        data = Reference(ws_monthly, min_col=2, min_row=1, max_row=len(df)+1)
        categories = Reference(ws_monthly, min_col=1, min_row=2, max_row=len(df)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 15
        chart.width = 25
        
        ws_monthly.add_chart(chart, "D2")
        
        # Save workbook
        wb.save(output_file)
        print(f"\n✅ Advanced report generated: {output_file}")
        print("📊 Includes data tables and professional charts")
        
        return output_file
    
    def export_data_for_analysis(self, output_dir='exported_data'):
        """Export multiple datasets as CSV for further analysis"""
        
        print("\n" + "="*70)
        print("📁 EXPORTING DATA TO CSV FILES")
        print("="*70)
        
        # Create output directory
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"✅ Created directory: {output_dir}")
        
        # Export 1: All orders with details
        print("\n📄 Exporting: complete_orders.csv...")
        query = '''
        SELECT 
            o.order_id,
            o.order_date,
            c.name as customer_name,
            c.city,
            c.customer_type,
            p.product_name,
            p.category,
            o.quantity,
            o.total_amount,
            o.status
        FROM orders o
        JOIN customers c ON o.customer_id = c.customer_id
        JOIN products p ON o.product_id = o.product_id;
        '''
        df = pd.read_sql_query(query, self.conn)
        df.to_csv(f'{output_dir}/complete_orders.csv', index=False)
        print(f"   ✅ Exported {len(df)} orders")
        
        # Export 2: Customer summary
        print("📄 Exporting: customer_summary.csv...")
        query = '''
        SELECT 
            c.customer_id,
            c.name,
            c.email,
            c.city,
            c.customer_type,
            c.join_date,
            COUNT(o.order_id) as total_orders,
            SUM(CASE WHEN o.status = 'Completed' THEN 1 ELSE 0 END) as completed_orders,
            ROUND(SUM(CASE WHEN o.status = 'Completed' THEN o.total_amount ELSE 0 END), 2) as total_spent
        FROM customers c
        LEFT JOIN orders o ON c.customer_id = o.customer_id
        GROUP BY c.customer_id, c.name, c.email, c.city, c.customer_type, c.join_date;
        '''
        df = pd.read_sql_query(query, self.conn)
        df.to_csv(f'{output_dir}/customer_summary.csv', index=False)
        print(f"   ✅ Exported {len(df)} customers")
        
        # Export 3: Product performance
        print("📄 Exporting: product_performance.csv...")
        query = '''
        SELECT 
            p.product_id,
            p.product_name,
            p.category,
            p.price,
            p.stock_quantity,
            COUNT(o.order_id) as times_ordered,
            SUM(o.quantity) as units_sold,
            ROUND(SUM(o.total_amount), 2) as total_revenue
        FROM products p
        LEFT JOIN orders o ON p.product_id = o.product_id AND o.status = 'Completed'
        GROUP BY p.product_id, p.product_name, p.category, p.price, p.stock_quantity;
        '''
        df = pd.read_sql_query(query, self.conn)
        df.to_csv(f'{output_dir}/product_performance.csv', index=False)
        print(f"   ✅ Exported {len(df)} products")
        
        print(f"\n✅ All data exported to: {output_dir}/")
        print("📊 Ready for Pandas analysis, visualization, or ML!")
    
    def close(self):
        """Close database connection"""
        self.conn.close()
        print("\n✅ Database connection closed")


def main():
    """Main execution"""
    print("="*70)
    print("🎯 DAY 16: EXCEL AUTOMATION PROJECT")
    print("="*70)
    print("\nAutomated Report Generation:")
    print("• SQL queries → Excel reports")
    print("• Professional formatting")
    print("• Charts and visualizations")
    print("• CSV exports for analysis")
    print("\n" + "="*70)
    
    # Initialize generator
    generator = ExcelReportGenerator()
    
    # Generate reports
    print("\n🔹 TASK 1: Basic Sales Summary")
    input("Press Enter to generate...")
    report1 = generator.create_sales_summary_report()
    
    print("\n🔹 TASK 2: Advanced Report with Charts")
    input("Press Enter to generate...")
    report2 = generator.create_advanced_report_with_charts()
    
    print("\n🔹 TASK 3: Export Data to CSV")
    input("Press Enter to export...")
    generator.export_data_for_analysis()
    
    # Close connection
    generator.close()
    
    print("\n" + "="*70)
    print("🎉 DAY 16 FILE AUTOMATION COMPLETE!")
    print("="*70)
    print("\nFiles created:")
    print(f"• {report1} - Multi-sheet sales summary")
    print(f"• {report2} - Report with charts")
    print("• exported_data/ - CSV files for analysis")
    
    print("\n💼 REAL-WORLD APPLICATION:")
    print("This is what data analysts do:")
    print("1. Query database (SQL)")
    print("2. Generate Excel reports (openpyxl)")
    print("3. Share with stakeholders")
    print("4. Automate weekly/monthly")
    
    print("\n🎯 You can now:")
    print("✅ Combine SQL with Excel")
    print("✅ Create professional reports")
    print("✅ Add charts automatically")
    print("✅ Export data for analysis")
    print("✅ Automate repetitive tasks")


if __name__ == "__main__":
    main()
